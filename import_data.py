"""
Revenue Report — Importer
Reads `Final_Revenue_Mapping_Cursor.xlsx` and (re)builds `wfm_data.db`.

Tables created:
  • revenue_hcr   — full headcount roster (Revenue HCR sheet)
  • revenue_team  — unified per-manager team performance (7 manager tabs)
  • revenue_meta  — small KV store (e.g. last_loaded_at)

Drops the old WFM tables (cost_summary, dadk_*, productivity_*, adara_*,
devops_*, it_helpdesk*, soho_*, customer_experience*) so the DB reflects
only the new structure.
"""
import os
import sqlite3
import datetime as dt
from typing import Any

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(HERE, "Final_Revenue_Mapping_Cursor.xlsx")
GRR_NRR_XLSX = os.path.join(HERE, "GRR_NRR_Account_Analysis.xlsx")
LEADER_PERF_XLSX = os.path.join(HERE, "Rev_Perf_Leader.xlsx")
DB_PATH = os.path.join(HERE, "wfm_data.db")

# Source-file leader names (sometimes include middle names / spelling variants)
# → canonical names used across the dashboard.
LEADER_NAME_MAP = {
    "Anurag Vinod Jain":             "Anurag Jain",
    "Anurag Jain":                   "Anurag Jain",
    "Keith Christopher Toby March":  "Toby March",
    "Toby March":                    "Toby March",
    "Sanchit Garg":                  "Sanchit Garg",
    "Ashish Sikka":                  "Ashish Sikka",
    "Vinay Varma":                   "Vinay Verma",   # spelling variant
    "Vinay Verma":                   "Vinay Verma",
    "Humberto Bifani":               "Humberto Bifani",
    "Humberto L Bifani":             "Humberto Bifani",
    "Carla Sue Shaw":                "Carla Shaw",
    "Carla Shaw":                    "Carla Shaw",
    # Newcomers from Rev_Perf_Leader.xlsx (regions / additional managers)
    "Yogeesh Chandra":               "Yogeesh Chandra",
    "EUROPE":                        "Ashish Sikka",   # EUROPE region rolls up to Ashish Sikka
}

MANAGER_TABS = [
    "Anurag Jain",
    "Carla Shaw",
    "Ashish Sikka",
    "Humberto Bifani",
    "Sanchit Garg",
    "Toby March",
    "Vinay Verma",
]

OLD_TABLES = [
    "cost_summary",
    "dadk_ctc",
    "dadk_headcount",
    "dadk_new_joiners",
    "productivity_emp",
    "adara_devops",
    "devops_uptime",
    "devops_tickets",
    "it_helpdesk",
    "it_helpdesk_notes",
    "soho_monitoring",
    "soho_client_success",
    "soho_social_content",
    "customer_experience",
    "customer_experience_notes",
]


# ─────────────────────────── helpers ───────────────────────────
def to_num(v: Any):
    """Convert numeric-ish to float; preserve None for blanks/'-'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s == "-" or s.lower() == "nan":
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def to_str(v: Any):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    return s if s else None


def is_total_row(team_val: str | None, emp_id_val: Any) -> bool:
    """Detect subtotal / grand-total rows (no Emp Id, Team has 'Total')."""
    if team_val is None:
        return False
    t = str(team_val).strip().lower()
    if t.endswith("total") or t == "grand total":
        return True
    # Some sheets put empty Emp Id rows that aren't totals — keep those as data
    return False


# ─────────────────────────── schema ───────────────────────────
DDL_HCR = """
CREATE TABLE revenue_hcr (
    row_order            INTEGER PRIMARY KEY,
    employee_id          TEXT,
    full_name            TEXT,
    doj                  TEXT,
    tenure               TEXT,
    official_email       TEXT,
    hrbp_name            TEXT,
    status               TEXT,
    gender               TEXT,
    leader               TEXT,
    emp_type             TEXT,
    manager_name         TEXT,
    entity               TEXT,
    division             TEXT,
    sub_division         TEXT,
    department           TEXT,
    sub_department       TEXT,
    designation          TEXT,
    office_location      TEXT,
    direct_manager_email TEXT,
    contribution_level   TEXT,
    date_of_exit         TEXT,
    employee_subtype     TEXT,
    band                 TEXT,
    q4_remarks_hr        TEXT,
    q4_remarks           TEXT,
    q3_remarks_aditi     TEXT
);
"""

DDL_TEAM = """
CREATE TABLE revenue_team (
    row_order               INTEGER PRIMARY KEY AUTOINCREMENT,
    manager_tab             TEXT NOT NULL,    -- "Anurag Jain", etc.
    team                    TEXT,
    status                  TEXT,             -- Active / Inactive / Terminated / (subtotal)
    emp_id                  TEXT,
    emp_name                TEXT,
    tenure_ymd              TEXT,
    budget_fy_25_26         REAL,
    budget_ytd_25_26        REAL,
    new_sales_25_26         REAL,
    ach_pct_25_26           REAL,
    salary_25_26            REAL,
    salary_multiple_25_26   REAL,
    total_expenses_25_26    REAL,
    sales_multiple_25_26    REAL,
    grr                     REAL,
    nrr                     REAL,
    q4_pipe_target          REAL,
    q4_pipe_creation        REAL,
    q4_pipe_achievement_pct REAL,
    q3_remarks              TEXT,
    q4_remarks              TEXT,
    is_total                INTEGER NOT NULL DEFAULT 0,  -- 0 = data, 1 = team subtotal, 2 = grand total
    sort_order              INTEGER NOT NULL DEFAULT 0
);
"""

DDL_META = """
CREATE TABLE revenue_meta (
    key   TEXT PRIMARY KEY,
    value TEXT
);
"""

DDL_ACCOUNTS = """
CREATE TABLE account_analysis (
    row_order        INTEGER PRIMARY KEY AUTOINCREMENT,
    account          TEXT,
    product          TEXT,
    am               TEXT,
    rev_24_25        REAL,
    churn            REAL,
    grr              REAL,
    downsell         REAL,
    upsell           REAL,
    nrr              REAL,
    new_revenue      REAL,
    rev_25_26        REAL
);
"""

DDL_LEADER_PERF = """
CREATE TABLE leader_perf_pivot (
    row_order            INTEGER PRIMARY KEY AUTOINCREMENT,
    leader_raw           TEXT,
    leader               TEXT,           -- canonical name from LEADER_NAME_MAP
    team                 TEXT,
    is_leader_total      INTEGER NOT NULL DEFAULT 0,  -- 1 for "{Leader} Total" rows
    is_grand_total       INTEGER NOT NULL DEFAULT 0,
    budget_fy_25_26      REAL,
    budget_ytd_25_26     REAL,
    new_sales_25_26      REAL,
    ach_pct_25_26        REAL,
    salary_25_26         REAL,
    salary_mult_25_26    REAL,
    commission_25_26     REAL,
    travel_exp_25_26     REAL,
    total_expenses_25_26 REAL,
    sales_mult_25_26     REAL,
    ach_pct_24_25        REAL,
    salary_mult_24_25    REAL,
    sales_mult_24_25     REAL
);
"""


# ─────────────────────────── importers ───────────────────────────
def import_hcr(ws, conn):
    """Import the 'Revenue HCR' sheet."""
    cur = conn.cursor()
    inserted = 0

    def cell(row, i):
        return row[i] if i < len(row) else None

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip fully empty rows
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        cur.execute(
            """INSERT INTO revenue_hcr (
                row_order, employee_id, full_name, doj, tenure, official_email,
                hrbp_name, status, gender, leader, emp_type, manager_name,
                entity, division, sub_division, department, sub_department,
                designation, office_location, direct_manager_email, contribution_level,
                date_of_exit, employee_subtype, band, q4_remarks_hr, q4_remarks,
                q3_remarks_aditi
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                idx,
                to_str(cell(row, 0)),
                to_str(cell(row, 1)),
                to_str(cell(row, 2)),
                to_str(cell(row, 3)),
                to_str(cell(row, 4)),
                to_str(cell(row, 5)),
                to_str(cell(row, 6)),
                to_str(cell(row, 7)),
                to_str(cell(row, 8)),
                to_str(cell(row, 9)),
                to_str(cell(row, 10)),
                to_str(cell(row, 11)),
                to_str(cell(row, 12)),
                to_str(cell(row, 13)),
                to_str(cell(row, 14)),
                to_str(cell(row, 15)),
                to_str(cell(row, 16)),
                to_str(cell(row, 17)),
                to_str(cell(row, 18)),
                to_str(cell(row, 19)),
                to_str(cell(row, 20)),
                to_str(cell(row, 21)),
                to_str(cell(row, 22)),
                to_str(cell(row, 23)),
                to_str(cell(row, 24)),
                to_str(cell(row, 25)),
            ),
        )
        inserted += 1
    print(f"  ✓ revenue_hcr: {inserted} rows")


def find_col(headers: list[str], *needles: str) -> int | None:
    """Locate a column index whose header matches any needle (case-insensitive substring)."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        hl = str(h).strip().lower().replace("\n", " ")
        for n in needles:
            if n.lower() in hl:
                return i
    return None


def import_leader_perf(conn):
    """Import the per-leader pivot (Rev_Perf_Leader.xlsx) — drives the Leaderboard view."""
    if not os.path.exists(LEADER_PERF_XLSX):
        print(f"  ⚠ {LEADER_PERF_XLSX} not found, skipping leader_perf_pivot import")
        return
    wb = openpyxl.load_workbook(LEADER_PERF_XLSX, data_only=True)
    ws = wb.active
    cur = conn.cursor()
    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        leader_raw = to_str(row[0])
        is_grand  = leader_raw and leader_raw.lower() == "grand total"
        is_total  = bool(leader_raw and leader_raw.endswith(" Total") and not is_grand)
        # For totals, the actual leader name strips the trailing " Total"
        if is_total:
            leader_name_only = leader_raw[:-6].strip()
        elif is_grand:
            leader_name_only = None
        else:
            leader_name_only = leader_raw
        leader_canonical = LEADER_NAME_MAP.get(leader_name_only, leader_name_only) if leader_name_only else None
        cur.execute(
            """INSERT INTO leader_perf_pivot (
                leader_raw, leader, team, is_leader_total, is_grand_total,
                budget_fy_25_26, budget_ytd_25_26, new_sales_25_26, ach_pct_25_26,
                salary_25_26, salary_mult_25_26, commission_25_26, travel_exp_25_26,
                total_expenses_25_26, sales_mult_25_26,
                ach_pct_24_25, salary_mult_24_25, sales_mult_24_25
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                leader_raw, leader_canonical, to_str(row[1]),
                1 if is_total else 0, 1 if is_grand else 0,
                to_num(row[2]),  to_num(row[3]),  to_num(row[4]),  to_num(row[5]),
                to_num(row[6]),  to_num(row[7]),  to_num(row[8]),  to_num(row[9]),
                to_num(row[10]), to_num(row[11]),
                to_num(row[12]), to_num(row[13]), to_num(row[14]),
            ),
        )
        inserted += 1
    print(f"  ✓ leader_perf_pivot: {inserted} rows")


def import_account_analysis(conn):
    """Import the GRR/NRR account-level analysis Excel."""
    if not os.path.exists(GRR_NRR_XLSX):
        print(f"  ⚠ {GRR_NRR_XLSX} not found, skipping account_analysis import")
        return
    wb = openpyxl.load_workbook(GRR_NRR_XLSX, data_only=True)
    sheet = "Export" if "Export" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    cur = conn.cursor()
    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        cur.execute(
            """INSERT INTO account_analysis (
                account, product, am, rev_24_25, churn, grr,
                downsell, upsell, nrr, new_revenue, rev_25_26
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (
                to_str(row[0]),  to_str(row[1]),  to_str(row[2]),
                to_num(row[3]),  to_num(row[4]),  to_num(row[5]),
                to_num(row[6]),  to_num(row[7]),  to_num(row[8]),
                to_num(row[9]),  to_num(row[10]),
            ),
        )
        inserted += 1
    print(f"  ✓ account_analysis: {inserted} accounts")


def import_manager(ws, manager_tab: str, conn):
    """Import a manager tab into revenue_team."""
    cur = conn.cursor()

    # Read the (multi-line) header row
    raw_headers = [c.value for c in ws[1]]

    # Map column positions — header text varies slightly across sheets
    col = {
        "team":   0,
        "status": 1,
        "empid":  2,
        "name":   3,
        "tenure": 4,
        "budget_fy":          find_col(raw_headers, "Budget FY"),
        "budget_ytd":         find_col(raw_headers, "Budget YTD"),
        "new_sales":          find_col(raw_headers, "New Sales"),
        "ach_25":             find_col(raw_headers, "Ach % (25-26)", "Ach% (25-26)"),
        "salary_25":          find_col(raw_headers, "Salary\nFY 25-26", "Salary FY"),
        "sal_mult_25":        find_col(raw_headers, "Salary Multiple (25-26)"),
        "expenses":           find_col(raw_headers, "Total Expenses"),
        "sales_mult_25":      find_col(raw_headers, "Sales Multiple (25-26)"),
        "grr":                find_col(raw_headers, "GRR"),
        "nrr":                find_col(raw_headers, "NRR"),
        "q4_target":          find_col(raw_headers, "Q4 Pipe Target"),
        "q4_creation":        find_col(raw_headers, "Q4 Pipe Creation"),
        "q4_pipe_ach":        find_col(raw_headers, "Q4 Pipe Achievement"),
        "q3_remarks":         find_col(raw_headers, "Q3 Remarks"),
        "q4_remarks":         find_col(raw_headers, "Q4 Remarks (HR-calibrated)", "Q4 Remarks"),
    }

    def cell(row, key):
        i = col.get(key)
        if i is None or i >= len(row):
            return None
        return row[i]

    inserted = 0
    sort_order = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        sort_order += 1

        team_val = cell(row, "team")
        team_str = to_str(team_val)
        emp_id_val = cell(row, "empid")
        is_total = 0
        if team_str:
            tl = team_str.lower()
            if tl == "grand total":
                is_total = 2
            elif tl.endswith("total"):
                is_total = 1

        cur.execute(
            """INSERT INTO revenue_team (
                manager_tab, team, status, emp_id, emp_name, tenure_ymd,
                budget_fy_25_26, budget_ytd_25_26, new_sales_25_26, ach_pct_25_26,
                salary_25_26, salary_multiple_25_26, total_expenses_25_26, sales_multiple_25_26,
                grr, nrr, q4_pipe_target, q4_pipe_creation, q4_pipe_achievement_pct,
                q3_remarks, q4_remarks,
                is_total, sort_order
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                manager_tab,
                team_str,
                to_str(cell(row, "status")),
                to_str(emp_id_val),
                to_str(cell(row, "name")),
                to_str(cell(row, "tenure")),
                to_num(cell(row, "budget_fy")),
                to_num(cell(row, "budget_ytd")),
                to_num(cell(row, "new_sales")),
                to_num(cell(row, "ach_25")),
                to_num(cell(row, "salary_25")),
                to_num(cell(row, "sal_mult_25")),
                to_num(cell(row, "expenses")),
                to_num(cell(row, "sales_mult_25")),
                to_num(cell(row, "grr")),
                to_num(cell(row, "nrr")),
                to_num(cell(row, "q4_target")),
                to_num(cell(row, "q4_creation")),
                to_num(cell(row, "q4_pipe_ach")),
                to_str(cell(row, "q3_remarks")),
                to_str(cell(row, "q4_remarks")),
                is_total,
                sort_order,
            ),
        )
        inserted += 1
    print(f"  ✓ revenue_team [{manager_tab}]: {inserted} rows")


# ─────────────────────────── main ───────────────────────────
def main():
    print(f"Reading {XLSX_PATH} …")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print(f"Connecting to {DB_PATH} …")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Drop old WFM tables + drop any prior revenue tables so we start clean
    print("Dropping old tables …")
    for t in OLD_TABLES + ["revenue_hcr", "revenue_team", "revenue_meta",
                           "account_analysis", "leader_perf_pivot"]:
        cur.execute(f"DROP TABLE IF EXISTS {t}")

    # Create fresh schema
    print("Creating new schema …")
    cur.execute(DDL_HCR)
    cur.execute(DDL_TEAM)
    cur.execute(DDL_META)
    cur.execute(DDL_ACCOUNTS)
    cur.execute(DDL_LEADER_PERF)

    # Index helpers
    cur.execute("CREATE INDEX idx_team_manager ON revenue_team(manager_tab)")
    cur.execute("CREATE INDEX idx_team_status ON revenue_team(status)")
    cur.execute("CREATE INDEX idx_hcr_status ON revenue_hcr(status)")
    cur.execute("CREATE INDEX idx_hcr_manager ON revenue_hcr(manager_name)")

    # Import HCR
    print("\nImporting Revenue HCR …")
    if "Revenue HCR" not in wb.sheetnames:
        raise RuntimeError("Sheet 'Revenue HCR' not found")
    import_hcr(wb["Revenue HCR"], conn)

    # Import each manager tab
    print("\nImporting manager tabs …")
    for tab in MANAGER_TABS:
        if tab not in wb.sheetnames:
            print(f"  ⚠ Sheet '{tab}' not found, skipping.")
            continue
        import_manager(wb[tab], tab, conn)

    # Import GRR/NRR account-level analysis
    print("\nImporting account-level GRR/NRR analysis …")
    import_account_analysis(conn)

    cur.execute("CREATE INDEX idx_acc_am ON account_analysis(am)")
    cur.execute("CREATE INDEX idx_acc_product ON account_analysis(product)")

    # Import leader-pivot (drives Leaderboard view)
    print("\nImporting leader performance pivot …")
    import_leader_perf(conn)
    cur.execute("CREATE INDEX idx_lp_leader ON leader_perf_pivot(leader)")
    cur.execute("CREATE INDEX idx_lp_total ON leader_perf_pivot(is_leader_total)")

    cur.execute(
        "INSERT OR REPLACE INTO revenue_meta(key, value) VALUES (?, ?)",
        ("last_loaded_at", dt.datetime.now().isoformat(timespec="seconds")),
    )
    cur.execute(
        "INSERT OR REPLACE INTO revenue_meta(key, value) VALUES (?, ?)",
        ("source_file", os.path.basename(XLSX_PATH)),
    )

    conn.commit()
    conn.close()
    print("\n✅ Import complete.")


if __name__ == "__main__":
    main()
