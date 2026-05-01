"""
Revenue Report Dashboard — RateGain
Interactive dashboard for Revenue HCR + per-manager team performance.
Backed by local SQLite (wfm_data.db) populated from
`Final_Revenue_Mapping_Cursor.xlsx` via `import_data.py`.
"""

import io
import os
import sqlite3
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, jsonify, request, send_file, session, redirect, url_for
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get("RG_SECRET_KEY", "rategain-revenue-dashboard-FY25-26-secret-key")
# Password gate. Override at deploy time via env var.
#   DASHBOARD_PASSWORD=mypassword python3 app.py
DASHBOARD_PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "rategain2026")
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wfm_data.db")


# ─── Auth decorator ───
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("authed"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


@app.before_request
def gate():
    """Require password for everything except the login flow + static assets."""
    if request.endpoint in ("login", "logout", "static") or session.get("authed"):
        return
    return redirect(url_for("login", next=request.path))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        if request.form.get("password", "") == DASHBOARD_PASSWORD:
            session["authed"] = True
            session.permanent = True
            nxt = request.args.get("next") or url_for("index")
            return redirect(nxt)
        error = "Incorrect password — please try again."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.pop("authed", None)
    return redirect(url_for("login"))

MANAGER_TABS = [
    "Anurag Jain",
    "Carla Shaw",
    "Ashish Sikka",
    "Humberto Bifani",
    "Sanchit Garg",
    "Toby March",
    "Vinay Verma",
]


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


@app.route("/")
def index():
    return render_template("dashboard.html", manager_tabs=MANAGER_TABS)


# ─────────────────────────── Revenue HCR ───────────────────────────
@app.route("/api/hcr")
def api_hcr():
    conn = get_db()
    status = request.args.get("status", "")
    manager = request.args.get("manager", "")
    division = request.args.get("division", "")
    leader = request.args.get("leader", "")
    search = request.args.get("q", "")

    where = ["1=1"]
    params: list = []
    if status:
        where.append("status = ?")
        params.append(status)
    if manager:
        where.append("manager_name = ?")
        params.append(manager)
    if division:
        where.append("division = ?")
        params.append(division)
    if leader:
        where.append("leader = ?")
        params.append(leader)
    if search:
        where.append(
            "(LOWER(full_name) LIKE ? OR LOWER(employee_id) LIKE ? OR LOWER(designation) LIKE ?)"
        )
        like = f"%{search.lower()}%"
        params.extend([like, like, like])

    sql = f"SELECT * FROM revenue_hcr WHERE {' AND '.join(where)} ORDER BY full_name"
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify(rows)


@app.route("/api/hcr/filters")
def api_hcr_filters():
    """Distinct values for HCR filter dropdowns."""
    conn = get_db()

    def distinct(col: str) -> list:
        return [
            r[0]
            for r in conn.execute(
                f"SELECT DISTINCT {col} FROM revenue_hcr WHERE {col} IS NOT NULL AND TRIM({col}) != '' ORDER BY {col}"
            ).fetchall()
        ]

    out = {
        "statuses": distinct("status"),
        "managers": distinct("manager_name"),
        "divisions": distinct("division"),
        "sub_divisions": distinct("sub_division"),
        "leaders": distinct("leader"),
        "departments": distinct("department"),
        "entities": distinct("entity"),
        "locations": distinct("office_location"),
    }
    conn.close()
    return jsonify(out)


@app.route("/api/hcr/summary")
def api_hcr_summary():
    """High-level KPIs and breakdowns for the HCR overview."""
    conn = get_db()
    cur = conn.cursor()

    total = cur.execute("SELECT COUNT(*) FROM revenue_hcr").fetchone()[0]
    active = cur.execute(
        "SELECT COUNT(*) FROM revenue_hcr WHERE status = 'Active'"
    ).fetchone()[0]
    inactive = cur.execute(
        "SELECT COUNT(*) FROM revenue_hcr WHERE status = 'Inactive'"
    ).fetchone()[0]

    by_division = [
        {"label": r[0] or "—", "value": r[1]}
        for r in cur.execute(
            "SELECT division, COUNT(*) FROM revenue_hcr GROUP BY division ORDER BY 2 DESC"
        ).fetchall()
    ]
    by_leader = [
        {"label": r[0] or "—", "value": r[1]}
        for r in cur.execute(
            "SELECT leader, COUNT(*) FROM revenue_hcr WHERE status='Active' GROUP BY leader ORDER BY 2 DESC LIMIT 12"
        ).fetchall()
    ]
    by_dept = [
        {"label": r[0] or "—", "value": r[1]}
        for r in cur.execute(
            "SELECT department, COUNT(*) FROM revenue_hcr GROUP BY department ORDER BY 2 DESC"
        ).fetchall()
    ]
    by_emp_type = [
        {"label": r[0] or "—", "value": r[1]}
        for r in cur.execute(
            "SELECT emp_type, COUNT(*) FROM revenue_hcr GROUP BY emp_type ORDER BY 2 DESC"
        ).fetchall()
    ]
    by_location = [
        {"label": r[0] or "—", "value": r[1]}
        for r in cur.execute(
            "SELECT office_location, COUNT(*) FROM revenue_hcr GROUP BY office_location ORDER BY 2 DESC LIMIT 10"
        ).fetchall()
    ]

    conn.close()
    return jsonify(
        {
            "total": total,
            "active": active,
            "inactive": inactive,
            "by_division": by_division,
            "by_leader": by_leader,
            "by_department": by_dept,
            "by_emp_type": by_emp_type,
            "by_location": by_location,
        }
    )


# ─────────────────────────── Manager team tabs ───────────────────────────
@app.route("/api/team/<path:manager>")
def api_team(manager: str):
    """All rows for a given manager tab (in original order)."""
    if manager not in MANAGER_TABS:
        return jsonify({"error": f"Unknown manager '{manager}'"}), 404

    conn = get_db()
    status = request.args.get("status", "")
    team = request.args.get("team", "")

    where = ["manager_tab = ?"]
    params: list = [manager]
    if status:
        # When filtering by status, also keep subtotal/grand-total rows
        where.append("(status = ? OR is_total > 0)")
        params.append(status)
    if team:
        where.append("(team = ? OR is_total = 2)")  # always include grand total
        params.append(team)

    sql = (
        f"SELECT * FROM revenue_team WHERE {' AND '.join(where)} ORDER BY sort_order"
    )
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify(rows)


@app.route("/api/team/<path:manager>/summary")
def api_team_summary(manager: str):
    """KPI summary for a manager tab — pulled directly from the Grand Total row."""
    if manager not in MANAGER_TABS:
        return jsonify({"error": "Unknown manager"}), 404

    conn = get_db()
    cur = conn.cursor()

    # Grand total row (is_total = 2)
    grand = cur.execute(
        "SELECT * FROM revenue_team WHERE manager_tab = ? AND is_total = 2 LIMIT 1",
        (manager,),
    ).fetchone()
    grand = dict(grand) if grand else {}

    # Counts of data rows by status
    counts_rows = cur.execute(
        """SELECT status, COUNT(*) FROM revenue_team
           WHERE manager_tab = ? AND is_total = 0
           GROUP BY status""",
        (manager,),
    ).fetchall()
    counts = {r[0] or "Unknown": r[1] for r in counts_rows}

    # Team-level subtotals (is_total = 1) — for stacked chart
    teams = [
        dict(r)
        for r in cur.execute(
            """SELECT team, budget_fy_25_26, budget_ytd_25_26, new_sales_25_26,
                      ach_pct_25_26, salary_25_26, total_expenses_25_26,
                      sales_multiple_25_26
               FROM revenue_team
               WHERE manager_tab = ? AND is_total = 1
               ORDER BY sort_order""",
            (manager,),
        ).fetchall()
    ]

    # Top contributors by New Sales
    top_sales = [
        dict(r)
        for r in cur.execute(
            """SELECT emp_name, team, new_sales_25_26, ach_pct_25_26,
                      sales_multiple_25_26
               FROM revenue_team
               WHERE manager_tab = ? AND is_total = 0 AND new_sales_25_26 > 0
               ORDER BY new_sales_25_26 DESC LIMIT 5""",
            (manager,),
        ).fetchall()
    ]

    # Distinct teams for the team filter
    teams_list = [
        r[0]
        for r in cur.execute(
            """SELECT DISTINCT team FROM revenue_team
               WHERE manager_tab = ? AND is_total = 0 AND team IS NOT NULL
               ORDER BY team""",
            (manager,),
        ).fetchall()
    ]

    conn.close()
    return jsonify(
        {
            "manager": manager,
            "grand_total": grand,
            "counts": counts,
            "teams": teams,
            "teams_list": teams_list,
            "top_sales": top_sales,
        }
    )


# ─────────────────────────── meta ───────────────────────────
@app.route("/api/download/<path:manager>")
def api_download(manager: str):
    """Generate an .xlsx export of a leader's full team performance + Q4 commentary."""
    if manager not in MANAGER_TABS:
        return jsonify({"error": f"Unknown leader '{manager}'"}), 404

    conn = get_db()
    rows = conn.execute(
        """SELECT team, status, emp_id, emp_name, tenure_ymd,
                  budget_fy_25_26, budget_ytd_25_26, new_sales_25_26, ach_pct_25_26,
                  salary_25_26, salary_multiple_25_26, total_expenses_25_26,
                  sales_multiple_25_26, grr, nrr,
                  q4_pipe_target, q4_pipe_creation, q4_pipe_achievement_pct,
                  q3_remarks, q4_remarks, is_total
           FROM revenue_team
           WHERE manager_tab = ?
           ORDER BY sort_order""",
        (manager,),
    ).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = manager[:31]  # Excel limit on sheet name length

    headers = [
        "Team", "Status", "Emp ID", "Name", "Tenure",
        "Budget FY 25-26", "Budget YTD 25-26", "New Sales 25-26", "Ach % (25-26)",
        "Salary 25-26", "Salary Multiple (25-26)", "Total Expenses 25-26",
        "Sales Multiple (25-26)", "GRR", "NRR",
        "Q4 Pipe Target", "Q4 Pipe Creation", "Q4 Pipe Ach %",
        "Q3 Remarks", "Q4 Remarks (HR-calibrated)",
    ]

    # Title row
    ws.cell(row=1, column=1, value=f"{manager} — Team Performance & Q4 Commentary").font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="5C2DB8")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d %b %Y, %H:%M')} · {len(rows)} rows").font = Font(italic=True, size=10, color="6B6B6B")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # Header row
    HEADER_ROW = 4
    header_fill = PatternFill("solid", fgColor="2A2C54")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(left=Side(style="thin", color="DDDDDD"),
                    right=Side(style="thin", color="DDDDDD"),
                    top=Side(style="thin", color="DDDDDD"),
                    bottom=Side(style="thin", color="DDDDDD"))
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    # Data rows
    for r_idx, row in enumerate(rows, start=HEADER_ROW + 1):
        is_total = row["is_total"] or 0
        values = [
            row["team"], row["status"], row["emp_id"], row["emp_name"], row["tenure_ymd"],
            row["budget_fy_25_26"], row["budget_ytd_25_26"], row["new_sales_25_26"], row["ach_pct_25_26"],
            row["salary_25_26"], row["salary_multiple_25_26"], row["total_expenses_25_26"],
            row["sales_multiple_25_26"], row["grr"], row["nrr"],
            row["q4_pipe_target"], row["q4_pipe_creation"], row["q4_pipe_achievement_pct"],
            row["q3_remarks"], row["q4_remarks"],
        ]
        row_fill = None
        if is_total == 2:
            row_fill = PatternFill("solid", fgColor="3A1A6A"); font_color = "FFFFFF"; bold = True
        elif is_total == 1:
            row_fill = PatternFill("solid", fgColor="2C1E57"); font_color = "FFFFFF"; bold = True
        else:
            font_color = "1A1A1A"; bold = False
        for c_idx, v in enumerate(values, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            c.font = Font(color=font_color, bold=bold, size=10)
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(c_idx >= 19),  # Q3/Q4 remarks wrap
                                    horizontal="right" if c_idx >= 6 and c_idx <= 18 else "left")
            c.border = border
            if row_fill:
                c.fill = row_fill
            # Number / percent / currency formats
            if c_idx in (6, 7, 8, 10, 12, 16, 17):  # currency-ish
                c.number_format = "#,##0"
            elif c_idx in (9, 14, 15, 18):  # percent
                c.number_format = "0.0%"
            elif c_idx in (11, 13):  # multiples
                c.number_format = "0.00\"x\""

    # Column widths
    widths = [22, 12, 10, 28, 16, 16, 16, 16, 14, 14, 16, 16, 16, 10, 10, 16, 16, 14, 60, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[HEADER_ROW].height = 36
    ws.freeze_panes = "E5"  # freeze top headers + Team/Status/EmpId/Name columns

    # Stream the file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = manager.replace(" ", "_")
    filename = f"{safe_name}_Team_Performance_FY25-26.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/meta")
def api_meta():
    conn = get_db()
    rows = conn.execute("SELECT key, value FROM revenue_meta").fetchall()
    conn.close()
    return jsonify({r[0]: r[1] for r in rows})


@app.route("/api/leaderboard")
def api_leaderboard():
    """Cross-leader performance comparison sourced from the
    `leader_perf_pivot` table (Rev_Perf_Leader.xlsx).

    The pivot file already gives us the leader-Total rows with all key metrics
    (Budget, New Sales, Ach %, Salary Mult, Sales Mult — both 25-26 and 24-25),
    so the CEO sees the same numbers HR shipped in the source spreadsheet.
    """
    conn = get_db()
    cur = conn.cursor()

    # Per-leader rollup — pulled directly from the leader-Total rows in the pivot.
    # Yogeesh Chandra is excluded from the executive leaderboard view per CEO request.
    leaders = [dict(r) for r in cur.execute(
        """SELECT
              leader                                                  AS leader,
              budget_fy_25_26                                         AS budget_fy,
              new_sales_25_26                                         AS new_sales,
              ach_pct_25_26                                           AS comp_ach_pct,
              salary_25_26                                            AS salary,
              salary_mult_25_26                                       AS comp_salary_mult,
              total_expenses_25_26                                    AS expenses,
              sales_mult_25_26                                        AS comp_sales_mult,
              ach_pct_24_25                                           AS prev_ach_pct,
              salary_mult_24_25                                       AS prev_salary_mult,
              sales_mult_24_25                                        AS prev_sales_mult
           FROM leader_perf_pivot
           WHERE is_leader_total = 1 AND leader != 'Yogeesh Chandra'
           ORDER BY new_sales_25_26 DESC"""
    ).fetchall()]

    # Headcount (Active / Inactive) is still needed for context — pull from revenue_team where possible
    hc = {r["manager_tab"]: {"active": r["a"], "inactive": r["i"]} for r in cur.execute(
        """SELECT manager_tab,
                  SUM(CASE WHEN status = 'Active' THEN 1 ELSE 0 END) AS a,
                  SUM(CASE WHEN status != 'Active' OR status IS NULL THEN 1 ELSE 0 END) AS i
           FROM revenue_team WHERE is_total = 0
           GROUP BY manager_tab"""
    ).fetchall()}

    for L in leaders:
        h = hc.get(L["leader"], {"active": 0, "inactive": 0})
        L["active_hc"]    = h["active"]
        L["inactive_hc"]  = h["inactive"]
        L["sales_per_emp"] = (L["new_sales"] / L["active_hc"]) if L["active_hc"] else None

    def _rank(metric, desc=True):
        eligible = [L for L in leaders if L.get(metric) is not None]
        eligible.sort(key=lambda x: (x[metric] or 0), reverse=desc)
        return [{"leader": L["leader"], "value": L[metric]} for L in eligible]

    rankings = {
        "sales_total":      _rank("new_sales"),
        "comp_ach_pct":     _rank("comp_ach_pct"),
        "comp_sales_mult":  _rank("comp_sales_mult"),
        "comp_salary_mult": _rank("comp_salary_mult"),
        "budget_fy":        _rank("budget_fy"),
        "sales_per_emp":    _rank("sales_per_emp"),
    }

    # Top 10 individuals across the entire org by Sales Achievement %
    top_individuals = [dict(r) for r in cur.execute(
        """SELECT manager_tab AS leader, emp_name, team,
                  budget_fy_25_26, new_sales_25_26,
                  ach_pct_25_26, salary_multiple_25_26, sales_multiple_25_26,
                  tenure_ymd
           FROM revenue_team
           WHERE is_total = 0 AND status = 'Active'
             AND ach_pct_25_26 IS NOT NULL
             AND budget_fy_25_26 IS NOT NULL AND budget_fy_25_26 > 50000
           ORDER BY ach_pct_25_26 DESC LIMIT 10"""
    ).fetchall()]

    # Top 10 by absolute new sales (biggest revenue producers)
    top_by_sales = [dict(r) for r in cur.execute(
        """SELECT manager_tab AS leader, emp_name, team,
                  new_sales_25_26, ach_pct_25_26, salary_multiple_25_26
           FROM revenue_team
           WHERE is_total = 0 AND status = 'Active'
             AND new_sales_25_26 IS NOT NULL AND new_sales_25_26 > 0
           ORDER BY new_sales_25_26 DESC LIMIT 10"""
    ).fetchall()]

    # Org-wide totals — prefer the Grand Total row from the pivot file when present
    grand = cur.execute(
        """SELECT budget_fy_25_26, new_sales_25_26, ach_pct_25_26,
                  salary_25_26, salary_mult_25_26, total_expenses_25_26, sales_mult_25_26
           FROM leader_perf_pivot WHERE is_grand_total = 1 LIMIT 1"""
    ).fetchone()
    if grand:
        totals = {
            "budget_fy":        grand["budget_fy_25_26"]      or 0,
            "new_sales":        grand["new_sales_25_26"]      or 0,
            "comp_ach_pct":     grand["ach_pct_25_26"],
            "salary":           grand["salary_25_26"]         or 0,
            "comp_salary_mult": grand["salary_mult_25_26"],
            "expenses":         grand["total_expenses_25_26"] or 0,
            "comp_sales_mult":  grand["sales_mult_25_26"],
        }
    else:
        totals = {
            "budget_fy":  sum((L["budget_fy"] or 0) for L in leaders),
            "new_sales":  sum((L["new_sales"] or 0) for L in leaders),
            "salary":     sum((L["salary"]    or 0) for L in leaders),
            "expenses":   sum((L["expenses"]  or 0) for L in leaders),
            "comp_ach_pct": None, "comp_sales_mult": None, "comp_salary_mult": None,
        }
        totals["comp_ach_pct"] = (totals["new_sales"] / totals["budget_fy"]) if totals["budget_fy"] else None
    totals["active_hc"]    = sum(L["active_hc"]   for L in leaders)
    totals["inactive_hc"]  = sum(L["inactive_hc"] for L in leaders)
    totals["sales_per_emp"] = (totals["new_sales"] / totals["active_hc"]) if totals["active_hc"] else None

    conn.close()
    return jsonify({
        "leaders":         leaders,
        "rankings":        rankings,
        "top_individuals": top_individuals,
        "top_by_sales":    top_by_sales,
        "totals":          totals,
    })


@app.route("/api/grrnrr")
def api_grrnrr():
    """Account-level GRR/NRR analysis for the CEO view.

    Data conventions in source spreadsheet:
      - `churn`    is stored as a NEGATIVE number (loss).
      - `downsell` is stored as a NEGATIVE number (loss).
      - `upsell`, `new_revenue` are POSITIVE.
      - The sheet contains a "Total" row and an "Applied filters:" row with NULL
        product / NULL AM — these must be excluded from all aggregations.

    All SUM-of-loss values are flipped to absolute (positive) for display.
    """
    conn = get_db()
    cur  = conn.cursor()

    # Common filter — exclude Total / filter-info rows that have no product nor AM.
    BASE = "FROM account_analysis WHERE product IS NOT NULL AND TRIM(product) != ''"

    # --- Headline KPIs ---
    head = cur.execute(
        f"""SELECT
            COUNT(*)                                                            AS total_accounts,
            COALESCE(SUM(rev_24_25), 0)                                         AS rev_24_25,
            COALESCE(SUM(rev_25_26), 0)                                         AS rev_25_26,
            ABS(COALESCE(SUM(CASE WHEN churn    < 0 THEN churn    ELSE 0 END), 0)) AS total_churn,
            ABS(COALESCE(SUM(CASE WHEN downsell < 0 THEN downsell ELSE 0 END), 0)) AS total_downsell,
            COALESCE(SUM(CASE WHEN upsell      > 0 THEN upsell      ELSE 0 END), 0) AS total_upsell,
            COALESCE(SUM(CASE WHEN new_revenue > 0 THEN new_revenue ELSE 0 END), 0) AS total_new_revenue,
            SUM(CASE WHEN churn    IS NOT NULL AND churn    < 0 THEN 1 ELSE 0 END) AS churned_accounts,
            SUM(CASE WHEN downsell IS NOT NULL AND downsell < 0 THEN 1 ELSE 0 END) AS downsell_accounts,
            SUM(CASE WHEN upsell   IS NOT NULL AND upsell   > 0 THEN 1 ELSE 0 END) AS upsell_accounts,
            SUM(CASE WHEN nrr IS NOT NULL AND nrr > 1.10 THEN 1 ELSE 0 END)        AS growth_accounts,
            SUM(CASE WHEN nrr IS NOT NULL AND nrr < 0.90 THEN 1 ELSE 0 END)        AS at_risk_accounts,
            SUM(CASE WHEN grr IS NOT NULL AND grr = 0    THEN 1 ELSE 0 END)        AS fully_lost_accounts,
            SUM(CASE WHEN new_revenue IS NOT NULL AND new_revenue > 0 THEN 1 ELSE 0 END) AS new_logo_accounts
           {BASE}"""
    ).fetchone()
    head = dict(head)

    # --- Composite (revenue-weighted) GRR / NRR ---
    weighted = cur.execute(
        f"""SELECT
            COALESCE(SUM(grr * rev_24_25) / NULLIF(SUM(rev_24_25), 0), 0) AS composite_grr,
            COALESCE(SUM(nrr * rev_24_25) / NULLIF(SUM(rev_24_25), 0), 0) AS composite_nrr
           {BASE} AND rev_24_25 IS NOT NULL AND rev_24_25 > 0"""
    ).fetchone()
    head["composite_grr"] = weighted["composite_grr"]
    head["composite_nrr"] = weighted["composite_nrr"]
    head["yoy_delta"] = (head["rev_25_26"] or 0) - (head["rev_24_25"] or 0)
    head["yoy_pct"]   = head["yoy_delta"] / head["rev_24_25"] if head["rev_24_25"] else 0

    # --- Product cohort breakdown ---
    products = [dict(r) for r in cur.execute(
        f"""SELECT
            product                                                                  AS product,
            COUNT(*)                                                                 AS accounts,
            COALESCE(SUM(rev_24_25), 0)                                              AS rev_24_25,
            COALESCE(SUM(rev_25_26), 0)                                              AS rev_25_26,
            ABS(COALESCE(SUM(CASE WHEN churn < 0 THEN churn ELSE 0 END), 0))         AS churn,
            COALESCE(SUM(CASE WHEN upsell > 0 THEN upsell ELSE 0 END), 0)            AS upsell,
            ABS(COALESCE(SUM(CASE WHEN downsell < 0 THEN downsell ELSE 0 END), 0))   AS downsell,
            COALESCE(SUM(CASE WHEN new_revenue > 0 THEN new_revenue ELSE 0 END), 0)  AS new_revenue,
            COALESCE(SUM(grr * rev_24_25) / NULLIF(SUM(rev_24_25), 0), NULL)         AS grr,
            COALESCE(SUM(nrr * rev_24_25) / NULLIF(SUM(rev_24_25), 0), NULL)         AS nrr
           {BASE}
           GROUP BY product
           ORDER BY rev_25_26 DESC"""
    ).fetchall()]

    # --- AM leaderboard (only AMs with revenue book) ---
    ams = [dict(r) for r in cur.execute(
        f"""SELECT
            am                                                                       AS am,
            COUNT(*)                                                                 AS accounts,
            COALESCE(SUM(rev_24_25), 0)                                              AS rev_24_25,
            COALESCE(SUM(rev_25_26), 0)                                              AS rev_25_26,
            ABS(COALESCE(SUM(CASE WHEN churn < 0 THEN churn ELSE 0 END), 0))         AS churn,
            COALESCE(SUM(CASE WHEN upsell > 0 THEN upsell ELSE 0 END), 0)            AS upsell,
            ABS(COALESCE(SUM(CASE WHEN downsell < 0 THEN downsell ELSE 0 END), 0))   AS downsell,
            COALESCE(SUM(CASE WHEN new_revenue > 0 THEN new_revenue ELSE 0 END), 0)  AS new_revenue,
            COALESCE(SUM(grr * rev_24_25) / NULLIF(SUM(rev_24_25), 0), NULL)         AS grr,
            COALESCE(SUM(nrr * rev_24_25) / NULLIF(SUM(rev_24_25), 0), NULL)         AS nrr
           {BASE}
             AND am IS NOT NULL AND TRIM(am) != ''
           GROUP BY am
           HAVING rev_24_25 > 0
           ORDER BY rev_25_26 DESC"""
    ).fetchall()]

    # --- Top accounts by 25-26 revenue (crown jewels) ---
    top_revenue = [dict(r) for r in cur.execute(
        f"""SELECT account, product, am, rev_24_25, rev_25_26, grr, nrr
           {BASE} AND rev_25_26 IS NOT NULL
           ORDER BY rev_25_26 DESC LIMIT 10"""
    ).fetchall()]

    # --- Biggest churn losses (red flags). Source values are negative — display as positive losses. ---
    top_churn = [dict(r) for r in cur.execute(
        f"""SELECT account, product, am, rev_24_25, rev_25_26,
                  ABS(churn) AS churn, grr, nrr
           {BASE} AND churn IS NOT NULL AND churn < 0
           ORDER BY ABS(churn) DESC LIMIT 10"""
    ).fetchall()]

    # --- Biggest upsell wins ---
    top_upsell = [dict(r) for r in cur.execute(
        f"""SELECT account, product, am, rev_24_25, rev_25_26, upsell, nrr
           {BASE} AND upsell IS NOT NULL AND upsell > 0
           ORDER BY upsell DESC LIMIT 10"""
    ).fetchall()]

    # --- At-risk accounts (NRR < 90%, revenue > $50K, not fully churned) ---
    at_risk = [dict(r) for r in cur.execute(
        f"""SELECT account, product, am, rev_24_25, rev_25_26,
                  ABS(COALESCE(churn, 0))    AS churn,
                  ABS(COALESCE(downsell, 0)) AS downsell,
                  grr, nrr
           {BASE} AND nrr IS NOT NULL AND nrr < 0.90
             AND rev_24_25 > 50000
             AND (grr IS NULL OR grr > 0)
           ORDER BY rev_24_25 DESC LIMIT 15"""
    ).fetchall()]

    conn.close()
    return jsonify({
        "summary":     head,
        "products":    products,
        "ams":         ams,
        "top_revenue": top_revenue,
        "top_churn":   top_churn,
        "top_upsell":  top_upsell,
        "at_risk":     at_risk,
    })


@app.route("/api/team_counts")
def api_team_counts():
    """Active vs Inactive HC per manager — drives the tab pill badges."""
    conn = get_db()
    rows = conn.execute(
        """SELECT manager_tab,
                  SUM(CASE WHEN status = 'Active' THEN 1 ELSE 0 END) AS active,
                  SUM(CASE WHEN status != 'Active' OR status IS NULL THEN 1 ELSE 0 END) AS inactive
           FROM revenue_team
           WHERE is_total = 0
           GROUP BY manager_tab"""
    ).fetchall()
    conn.close()
    return jsonify({r["manager_tab"]: {"active": r["active"], "inactive": r["inactive"]} for r in rows})


if __name__ == "__main__":
    print("🚀 Revenue Report Dashboard at http://127.0.0.1:5050")
    app.run(debug=True, host="127.0.0.1", port=5050)
